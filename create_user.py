# Mario Wenzl 5CN
import argparse
import logging
import os
import random
import string
from logging.handlers import RotatingFileHandler
import unicodedata
from openpyxl import load_workbook
from plistlib import InvalidFileException

"""
Argument Parser; Helptext erklärt Funktion
"""
parser = argparse.ArgumentParser()
parser.add_argument('file')
group = parser.add_mutually_exclusive_group(required=False)
group.add_argument("-v", "--verbosity", help="increase output verbosity", action="store_true")
group.add_argument("-q", "--quiet", help="Run with no output", action="store_true")

args = parser.parse_args()

"""
Aktionen werden in einem Logfile festgehalten
Logfile: logs_class.log
"""
logger = logging.getLogger("create users")
handler = RotatingFileHandler("logs_user.log", maxBytes=500000, backupCount=20)
streamhandler = logging.StreamHandler()
logger.addHandler(handler)
logger.addHandler(streamhandler)

infile = args.file

logger.setLevel(logging.INFO)
if args.quiet:
    logger.setLevel(logging.WARNING)
elif args.verbosity:
    logger.setLevel(logging.DEBUG)


def read_file(file, start_line, end_line):
    """
    Zeilen des ersten Excel-Sheets einlesen und als Generator zurück geben
    :param file: Pfad und Name des Excel Sheets, das Eingelesen werden soll
    :param start_line: Zeile, in welcher angefangen werden soll zu lesen
    :param end_line: Zeile, in welcher aufgehört werden soll zu lesen
    :return: Generator mit dem Inhalt der Zeilen
    """
    try:
        wb = load_workbook(file, read_only=True)
    except InvalidFileException:
        logger.warning(f"{file} is not supported")
        return

    if wb is None:
        logging.warning(f"{file} could not be read")
        return
    ws = wb[wb.sheetnames[0]]
    logger.debug(f"{file} geöffnet")
    for row in ws.iter_rows(min_row=start_line, max_row=end_line):
        yield {"vorname": row[0].value, "nachname": row[1].value, "group": row[2].value, "class": row[3].value}
    yield {"vorname": "lehrer", "nachname": "lehrer", "group": "teacher", "class": "lehrer"}
    yield {"vorname": "seminar", "nachname": "seminar", "group": "teacher", "class": "seminar"}


def shave_marks(txt):
    """
    Löscht Sondersymbole; z.B.: â -> a
    :param txt: Originaltext
    :return: Geänderter Text
    """
    norm_txt = unicodedata.normalize('NFD', txt)
    shaved = ''.join(c for c in norm_txt
                     if not unicodedata.combining(c))
    return unicodedata.normalize('NFC', shaved)


def create_user():
    """
    Erstellt scripts zum Erstellen und Löschen der User aus dem Excel Sheet
    Dokumentiert zusätzlich die Namen und Passwörter der User in einem Text File
    Benutzererstellungsscript: create_user.sh
    Benutzerlöschungsscript: delete_user.sh
    Nutzer und Passwörter: users.txt
    """
    file = read_file(infile, 2, 38)
    with open("create_user.sh", "w") as create_user, open("delete_user.sh", "w") as delete_user, open(
            "users.txt", "w") as users:
        logger.info("Files wurden geöffnet")
        create_user.write("#!/bin/bash\n")
        delete_user.write("#!/bin/bash\n")
        create_user.write("mkdir -p /home/lehrer\n")
        create_user.write("mkdir -p /home/schueler\n")
        logger.debug("Ersten Schreibzugriffe getätigt")

        userlist = []
        for line in file:
            str1 = ""
            password = str1.join([random.choice(f"{string.ascii_letters}!%&(),._-=^#") for i in range(10)])
            username = shave_marks(line["nachname"].lower())
            logger.debug(f"Passwort für {username} erstellt")

            counter = 1
            while True:
                if username + str(counter) not in userlist and username not in userlist:
                    userlist.append(username)
                    break
                if shave_marks(line["nachname"].lower()) in userlist or username + str(counter) in userlist:
                    username = shave_marks(line["nachname"].lower()) + str(counter)
                    counter += 1
            logger.debug("Zahl an User falls dieser mehrfach vorkommt")

            create_user.write(
                f"if [ \"$?\" -ne 0 ]; then echo 'User {username} already exists'; exit 1; fi\n")
            if line["group"] == "teacher":
                users.write(f"{username} {password} saved in /home/lehrer\n")
                create_user.write(
                    f'useradd -m -d /home/lehrer -c "{username}" --groups cdrom,plugdev,sambashare -s /bin/bash {username}\n')
            else:
                users.write(f"{username} {password} saved in /home/schueler\n")
                create_user.write(
                    f'useradd -m -d /home/schueler -c "{username}" --groups cdrom,plugdev,sambashare -s /bin/bash {username}\n')
            create_user.write(
                f"echo {username}:\'{password}\' | chpasswd\n")

            delete_user.write(f"userdel -rf {username}\n")
            logger.info(f"Einträge für {username} erstellt")
        logger.info("Erstellen und Schreiben von FIles fertig")


"""
Ausführung des Scripts und Änderung der Berechtigungen des Files
"""
create_user()
os.chmod("create_user.sh", 0o777)
os.chmod("delete_user.sh", 0o777)
