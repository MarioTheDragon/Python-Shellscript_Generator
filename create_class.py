# Mario Wenzl 5CN
import argparse
import logging
import os
import random
from logging.handlers import RotatingFileHandler
from plistlib import InvalidFileException

from openpyxl import load_workbook

"""
Argument Parser; Helptext erklärt Funktion
"""
parser = argparse.ArgumentParser()
parser.add_argument('file')
group = parser.add_mutually_exclusive_group(required=False)
group.add_argument("-v", "--verbosity", help="increase output verbosity", action="store_true")
group.add_argument("-q", "--quiet", help="Run with no output", action="store_true")

args = parser.parse_args()

"""
Aktionen werden in einem Logfile festgehalten
Logfile: logs_class.log
"""
logger = logging.getLogger("create ckass")
handler = RotatingFileHandler("logs_class.log", maxBytes=500000, backupCount=20)
streamhandler = logging.StreamHandler()
logger.addHandler(handler)
logger.addHandler(streamhandler)

infile = args.file

logger.setLevel(logging.INFO)
if args.quiet:
    logger.setLevel(logging.WARNING)
elif args.verbosity:
    logger.setLevel(logging.DEBUG)


def read_file(file, start_line, end_line):
    """
    Zeilen des ersten Excel-Sheets einlesen und als Generator zurück geben
    :param file: Pfad und Name des Excel Sheets, das Eingelesen werden soll
    :param start_line: Zeile, in welcher angefangen werden soll zu lesen
    :param end_line: Zeile, in welcher aufgehört werden soll zu lesen
    :return: Generator mit dem Inhalt der Zeilen
    """
    try:
        wb = load_workbook(file, read_only=True)
    except InvalidFileException:
        logger.warning(f"{file} is not supported")
        return

    if wb is None:
        logging.warning(f"{file} konnte nicht gelesen werden")
        return
    ws = wb[wb.sheetnames[0]]
    logger.debug(f"{file} geöffnet")
    for row in ws.iter_rows(min_row=start_line, max_row=end_line):
        yield row


def create_class():
    """
    Erstellt scripts zum Erstellen und Löschen der User aus dem Excel Sheet
    Dokumentiert zusätzlich die Namen und Passwörter der User in einem Text File
    Benutzererstellungsscript: create_classes.sh
    Benutzerlöschungsscript: delete_classes.sh
    Nutzer und Passwörter: classes.txt
    """
    file = read_file(infile, 4, 42)
    with open("create_classes.sh", "w") as create_classes, open("delete_classes.sh", "w") as delete_classes, open(
            "classes.txt", "w") as classes:
        logger.info("Files wurden geöffnet")
        create_classes.write("#!/bin/bash\n")
        delete_classes.write("#!/bin/bash\n")
        create_classes.write("mkdir -p /home/klassen\n")
        logger.debug("ersten zeilen in das File geschrieben")

        for line in file:
            password = f"{line[0].value.lower()}{random.choice('!%&(),._-=^#')}{line[1].value}" \
                       f"{random.choice('!%&(),._-=^#')}{line[2].value.lower()}{random.choice('!%&(),._-=^#')}"
            logger.debug(f"Passwort für {line[0].value.lower()} generiert")
            create_classes.write(
                f"if [ \"$?\" -ne 0 ]; then echo 'User {line[0].value.lower()} already exists'; exit 1; fi\n")
            create_classes.write(
                f'useradd -m -d /home/klassen -c "k{line[0].value.lower()}'
                f'" --groups cdrom,plugdev,sambashare -s /bin/bash k{line[0].value.lower()}\n')
            create_classes.write(
                f"echo k{line[0].value.lower()}:\'{password}\' | chpasswd\n")

            classes.write(f"k{line[0].value.lower()} {password} saved in /home/klassen\n")
            delete_classes.write(f"userdel -rf k{line[0].value.lower()}\n")
            logger.debug(f"{line[0].value.lower()} ist fertig")
        logger.info("Scripts wurden erstellt")


"""
Ausführung des Scripts und Änderung der Berechtigungen des Files
"""
create_class()
os.chmod("create_classes.sh", 0o777)
os.chmod("delete_classes.sh", 0o777)
